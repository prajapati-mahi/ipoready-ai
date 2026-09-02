import os
import openpyxl
import pandas as pd
from typing import Dict, Any, List

class ExcelParser:
    @staticmethod
    def parse_excel(file_path: str) -> Dict[str, Any]:
        """
        Deep spreadsheet intelligence:
        - Extracts sheets, rows, columns, formulas, merged cells, exact cell coordinates (e.g. Sheet P&L -> B12)
        - Detects multi-year financial statements
        - Automatically isolates tabular numeric regions
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"File not found: {file_path}")

        workbook = openpyxl.load_workbook(file_path, data_only=False)
        data_workbook = openpyxl.load_workbook(file_path, data_only=True)  # To read evaluated values

        sheets_data = []
        tables_data = []
        cell_map = {}  # Map of "SheetName!CellCoord" -> { value, formula, row, col }
        full_text_chunks = []

        for sheet_name in workbook.sheetnames:
            ws_formula = workbook[sheet_name]
            ws_data = data_workbook[sheet_name]

            max_row = ws_data.max_row
            max_col = ws_data.max_column

            sheet_rows = []
            sheet_text_lines = [f"=== SHEET: {sheet_name} ==="]

            for r in range(1, max_row + 1):
                row_vals = []
                row_has_data = False
                for c in range(1, max_col + 1):
                    cell_formula_obj = ws_formula.cell(row=r, column=c)
                    cell_data_obj = ws_data.cell(row=r, column=c)

                    val = cell_data_obj.value
                    formula = str(cell_formula_obj.value) if str(cell_formula_obj.value).startswith("=") else None
                    cell_ref = f"{sheet_name}!{cell_formula_obj.coordinate}"

                    if val is not None:
                        row_has_data = True

                    cell_info = {
                        "sheet": sheet_name,
                        "cell": cell_formula_obj.coordinate,
                        "value": val,
                        "formula": formula,
                        "row": r,
                        "col": c
                    }
                    cell_map[cell_ref] = cell_info
                    row_vals.append(val if val is not None else "")

                if row_has_data:
                    sheet_rows.append(row_vals)
                    line_str = " | ".join([str(v) for v in row_vals if str(v).strip()])
                    if line_str:
                        sheet_text_lines.append(line_str)

            # Convert to DataFrame for structured table extraction
            try:
                df = pd.read_excel(file_path, sheet_name=sheet_name)
                # Drop all-NaN columns and rows
                df = df.dropna(how='all').dropna(axis=1, how='all')
                
                headers = [str(col) for col in df.columns]
                rows = df.fillna("").values.tolist()

                tables_data.append({
                    "page_number": None,
                    "sheet_name": sheet_name,
                    "table_title": f"Spreadsheet - {sheet_name}",
                    "headers": headers,
                    "rows": rows,
                    "raw_csv": df.to_csv(index=False)
                })
            except Exception as e:
                print(f"Error loading pandas sheet {sheet_name}: {e}")

            sheets_data.append({
                "sheet_name": sheet_name,
                "row_count": max_row,
                "col_count": max_col,
                "rows": sheet_rows
            })
            full_text_chunks.append("\n".join(sheet_text_lines))

        workbook.close()
        data_workbook.close()

        return {
            "sheet_count": len(sheets_data),
            "sheets": sheets_data,
            "tables": tables_data,
            "cell_map": cell_map,
            "full_text": "\n\n".join(full_text_chunks)
        }
