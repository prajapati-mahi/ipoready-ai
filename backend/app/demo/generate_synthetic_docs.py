import os
import openpyxl
from openpyxl.styles import Font, PatternFill
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

class SyntheticDocumentGenerator:
    @classmethod
    def generate_all(cls, output_dir: str):
        os.makedirs(output_dir, exist_ok=True)
        pdf_annual_path = os.path.join(output_dir, "Acme_Tech_Annual_Report_FY24.pdf")
        pdf_pres_path = os.path.join(output_dir, "Acme_Tech_Investor_Presentation_FY24.pdf")
        xlsx_model_path = os.path.join(output_dir, "Acme_Tech_Financial_Model_FY22_FY24.xlsx")
        xlsx_cf_path = os.path.join(output_dir, "Acme_Tech_Cash_Flow_Statement.xlsx")

        cls.generate_annual_report_pdf(pdf_annual_path)
        cls.generate_investor_presentation_pdf(pdf_pres_path)
        cls.generate_financial_model_excel(xlsx_model_path)
        cls.generate_cash_flow_excel(xlsx_cf_path)

        return [pdf_annual_path, pdf_pres_path, xlsx_model_path, xlsx_cf_path]

    @staticmethod
    def generate_annual_report_pdf(file_path: str):
        doc = SimpleDocTemplate(file_path, pagesize=letter, leftMargin=40, rightMargin=40, topMargin=40, bottomMargin=40)
        styles = getSampleStyleSheet()
        elements = []

        title_style = ParagraphStyle('TitleStyle', parent=styles['Heading1'], fontSize=18, textColor=colors.HexColor('#0F172A'), spaceAfter=8)
        h2_style = ParagraphStyle('H2Style', parent=styles['Heading2'], fontSize=13, textColor=colors.HexColor('#1E293B'), spaceBefore=8, spaceAfter=4)
        body_style = ParagraphStyle('BodyStyle', parent=styles['Normal'], fontSize=9, textColor=colors.HexColor('#334155'), leading=13)

        elements.append(Paragraph("ACME TECHNOLOGIES PRIVATE LIMITED", title_style))
        elements.append(Paragraph("<b>Annual Report & Audited Financial Statements - FY2023-24</b>", h2_style))
        elements.append(Paragraph("CIN: U72200MH2018PTC308912 | Statutory Auditors: Deloitte Haskins & Sells LLP", body_style))
        elements.append(Spacer(1, 10))

        elements.append(Paragraph("STATEMENT OF PROFIT AND LOSS (in Crores)", h2_style))
        pnl_data = [
            ["Particulars", "FY2023-24", "FY2022-23", "YoY Growth %"],
            ["Revenue from Operations", "125.00", "100.00", "25.0%"],
            ["Cost of Goods & Services", "56.25", "46.00", "22.3%"],
            ["Gross Profit", "68.75", "54.00", "27.3%"],
            ["Employee Benefit Expenses", "25.00", "20.00", "25.0%"],
            ["Other Operating Expenses", "12.50", "10.00", "25.0%"],
            ["EBITDA", "31.25", "24.00", "30.2%"],
            ["Depreciation & Amortization", "4.00", "3.20", "25.0%"],
            ["EBIT (Operating Profit)", "27.25", "20.80", "31.0%"],
            ["Finance Costs (Interest)", "3.25", "2.50", "30.0%"],
            ["Profit Before Tax (PBT)", "24.00", "18.30", "31.1%"],
            ["Tax Expense", "5.25", "4.10", "28.0%"],
            ["Profit After Tax (PAT)", "18.75", "14.20", "32.0%"]
        ]
        t = Table(pnl_data, colWidths=[180, 80, 80, 80])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0F172A')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('ALIGN', (1,0), (-1,-1), 'RIGHT'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 10))

        elements.append(Paragraph("BALANCE SHEET & CASH FLOW HIGHLIGHTS (Page 74-86)", h2_style))
        bs_text = (
            "Total Assets for FY2024 stood at 150 Cr compared to 120 Cr in FY2023. Total Debt increased to 42 Cr from 30 Cr. "
            "Net Worth reached 85 Cr. Cash and Cash Equivalents stood at 18.5 Cr. "
            "Operating Cash Flow for FY2024 was 27.2 Cr (down from 40 Cr in FY2023 due to working capital expansion in trade receivables). "
            "Top 5 clients contribute 42% of revenue. Segment contributions: Enterprise SaaS (68%) and Cloud Infrastructure (32%)."
        )
        elements.append(Paragraph(bs_text, body_style))
        doc.build(elements)

    @staticmethod
    def generate_investor_presentation_pdf(file_path: str):
        doc = SimpleDocTemplate(file_path, pagesize=letter, leftMargin=40, rightMargin=40, topMargin=40, bottomMargin=40)
        styles = getSampleStyleSheet()
        elements = []

        title_style = ParagraphStyle('TitleStyle', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#1E293B'))
        body_style = ParagraphStyle('BodyStyle', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#334155'), leading=14)

        elements.append(Paragraph("ACME TECHNOLOGIES - INVESTOR PRESENTATION FY2024", title_style))
        elements.append(Paragraph("<b>Operational Highlights & Strategic Overview</b>", body_style))
        elements.append(Spacer(1, 10))

        pres_text = (
            "<b>Key Operational Metrics:</b><br/>"
            "- Operational Reported Revenue FY24: <b>128 Cr</b> (includes unbilled project retainers).<br/>"
            "- EBITDA FY24: <b>32.5 Cr</b>.<br/>"
            "- Engineering Team Headcount: 240 engineers across Pune and Bangalore.<br/>"
            "- Net Retention Rate (NRR): 124%. Enterprise Clients: 310 accounts.<br/>"
            "- Segment Mix: Enterprise SaaS 70%, Cloud Infrastructure 30%."
        )
        elements.append(Paragraph(pres_text, body_style))
        doc.build(elements)

    @staticmethod
    def generate_financial_model_excel(file_path: str):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "P&L"

        header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")

        headers = ["Line Item", "FY2022 (in Cr)", "FY2023 (in Cr)", "FY2024 (in Cr)", "YoY Growth %"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill

        rows = [
            ("Revenue", 78.00, 100.00, 125.00, 25.0),
            ("COGS", 35.00, 46.00, 56.25, 22.3),
            ("Gross Profit", 43.00, 54.00, 68.75, 27.3),
            ("Opex", 26.00, 30.00, 37.50, 25.0),
            ("EBITDA", 17.00, 24.00, 31.25, 30.2),
            ("D&A", 2.50, 3.20, 4.00, 25.0),
            ("EBIT", 14.50, 20.80, 27.25, 31.0),
            ("Interest", 2.00, 2.50, 3.25, 30.0),
            ("PBT", 12.50, 18.30, 24.00, 31.1),
            ("Tax", 3.00, 4.10, 5.25, 28.0),
            ("PAT", 9.50, 14.20, 18.75, 32.0)
        ]

        for r_idx, row in enumerate(rows, 2):
            for c_idx, val in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=val)

        ws2 = wb.create_sheet(title="Balance Sheet")
        bs_headers = ["Particulars", "FY2023", "FY2024"]
        for col_idx, h in enumerate(bs_headers, 1):
            cell = ws2.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill

        bs_rows = [
            ("Total Assets", 120.00, 150.00),
            ("Total Liabilities", 52.00, 65.00),
            ("Total Debt", 30.00, 42.00),
            ("Net Worth", 68.00, 85.00),
            ("Cash & Equivalents", 14.00, 18.50)
        ]
        for r_idx, row in enumerate(bs_rows, 2):
            for c_idx, val in enumerate(row, 1):
                ws2.cell(row=r_idx, column=c_idx, value=val)

        wb.save(file_path)

    @staticmethod
    def generate_cash_flow_excel(file_path: str):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cash Flow"

        headers = ["Cash Flow Items", "FY2023 (in Cr)", "FY2024 (in Cr)"]
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=h)

        rows = [
            ("Operating Cash Flow (OCF)", 40.00, 27.20),
            ("Capital Expenditure (CapEx)", 10.00, 12.00),
            ("Free Cash Flow (FCF)", 30.00, 15.20),
            ("Financing Cash Flow", -5.00, 8.00),
            ("Net Change in Cash", 15.00, 4.50)
        ]
        for r_idx, row in enumerate(rows, 2):
            for c_idx, val in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=val)

        wb.save(file_path)
