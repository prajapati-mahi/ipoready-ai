# -*- coding: utf-8 -*-
import os

app_root = r"C:\Users\mahii\.gemini\antigravity\scratch\ipoready-ai\backend\app"

def write_f(rel_path, code):
    p = os.path.join(app_root, rel_path)
    os.makedirs(os.path.dirname(p), exist_ok=True)
    with open(p, "w", encoding="utf-8") as f:
        f.write(code.strip() + "\n")
    print(f"Created: {rel_path}")

# 1. Synthetic Document Generator
write_f("demo/generate_synthetic_docs.py", """
import os
import openpyxl
from openpyxl.styles import Font, PatternFill
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

class SyntheticDocumentGenerator:
    @classmethod
    def generate_all(cls, output_dir: str):
        os.makedirs(output_dir, exist_ok=True)
        pdf_annual_path = os.path.join(output_dir, "Acme_Tech_Annual_Report_FY24.pdf")
        pdf_pres_path = os.path.join(output_dir, "Acme_Tech_Investor_Presentation_FY24.pdf")
        xlsx_model_path = os.path.join(output_dir, "Acme_Tech_Financial_Model_FY22_FY24.xlsx")
        xlsx_cf_path = os.path.join(output_dir, "Acme_Tech_Cash_Flow_Statement.xlsx")

        cls.generate_annual_report_pdf(pdf_annual_path)
        cls.generate_investor_presentation_pdf(pdf_pres_path)
        cls.generate_financial_model_excel(xlsx_model_path)
        cls.generate_cash_flow_excel(xlsx_cf_path)

        return [pdf_annual_path, pdf_pres_path, xlsx_model_path, xlsx_cf_path]

    @staticmethod
    def generate_annual_report_pdf(file_path: str):
        doc = SimpleDocTemplate(file_path, pagesize=letter, leftMargin=40, rightMargin=40, topMargin=40, bottomMargin=40)
        styles = getSampleStyleSheet()
        elements = []

        title_style = ParagraphStyle('TitleStyle', parent=styles['Heading1'], fontSize=18, textColor=colors.HexColor('#0F172A'), spaceAfter=8)
        h2_style = ParagraphStyle('H2Style', parent=styles['Heading2'], fontSize=13, textColor=colors.HexColor('#1E293B'), spaceBefore=8, spaceAfter=4)
        body_style = ParagraphStyle('BodyStyle', parent=styles['Normal'], fontSize=9, textColor=colors.HexColor('#334155'), leading=13)

        elements.append(Paragraph("ACME TECHNOLOGIES PRIVATE LIMITED", title_style))
        elements.append(Paragraph("<b>Annual Report & Audited Financial Statements - FY2023-24</b>", h2_style))
        elements.append(Paragraph("CIN: U72200MH2018PTC308912 | Statutory Auditors: Deloitte Haskins & Sells LLP", body_style))
        elements.append(Spacer(1, 10))

        elements.append(Paragraph("STATEMENT OF PROFIT AND LOSS (in Crores)", h2_style))
        pnl_data = [
            ["Particulars", "FY2023-24", "FY2022-23", "YoY Growth %"],
            ["Revenue from Operations", "125.00", "100.00", "25.0%"],
            ["Cost of Goods & Services", "56.25", "46.00", "22.3%"],
            ["Gross Profit", "68.75", "54.00", "27.3%"],
            ["Employee Benefit Expenses", "25.00", "20.00", "25.0%"],
            ["Other Operating Expenses", "12.50", "10.00", "25.0%"],
            ["EBITDA", "31.25", "24.00", "30.2%"],
            ["Depreciation & Amortization", "4.00", "3.20", "25.0%"],
            ["EBIT (Operating Profit)", "27.25", "20.80", "31.0%"],
            ["Finance Costs (Interest)", "3.25", "2.50", "30.0%"],
            ["Profit Before Tax (PBT)", "24.00", "18.30", "31.1%"],
            ["Tax Expense", "5.25", "4.10", "28.0%"],
            ["Profit After Tax (PAT)", "18.75", "14.20", "32.0%"]
        ]
        t = Table(pnl_data, colWidths=[180, 80, 80, 80])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#0F172A')),
            ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('ALIGN', (1,0), (-1,-1), 'RIGHT'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#CBD5E1')),
        ]))
        elements.append(t)
        elements.append(Spacer(1, 10))

        elements.append(Paragraph("BALANCE SHEET & CASH FLOW HIGHLIGHTS (Page 74-86)", h2_style))
        bs_text = (
            "Total Assets for FY2024 stood at 150 Cr compared to 120 Cr in FY2023. Total Debt increased to 42 Cr from 30 Cr. "
            "Net Worth reached 85 Cr. Cash and Cash Equivalents stood at 18.5 Cr. "
            "Operating Cash Flow for FY2024 was 27.2 Cr (down from 40 Cr in FY2023 due to working capital expansion in trade receivables). "
            "Top 5 clients contribute 42% of revenue. Segment contributions: Enterprise SaaS (68%) and Cloud Infrastructure (32%)."
        )
        elements.append(Paragraph(bs_text, body_style))
        doc.build(elements)

    @staticmethod
    def generate_investor_presentation_pdf(file_path: str):
        doc = SimpleDocTemplate(file_path, pagesize=letter, leftMargin=40, rightMargin=40, topMargin=40, bottomMargin=40)
        styles = getSampleStyleSheet()
        elements = []

        title_style = ParagraphStyle('TitleStyle', parent=styles['Heading1'], fontSize=16, textColor=colors.HexColor('#1E293B'))
        body_style = ParagraphStyle('BodyStyle', parent=styles['Normal'], fontSize=10, textColor=colors.HexColor('#334155'), leading=14)

        elements.append(Paragraph("ACME TECHNOLOGIES - INVESTOR PRESENTATION FY2024", title_style))
        elements.append(Paragraph("<b>Operational Highlights & Strategic Overview</b>", body_style))
        elements.append(Spacer(1, 10))

        pres_text = (
            "<b>Key Operational Metrics:</b><br/>"
            "- Operational Reported Revenue FY24: <b>128 Cr</b> (includes unbilled project retainers).<br/>"
            "- EBITDA FY24: <b>32.5 Cr</b>.<br/>"
            "- Engineering Team Headcount: 240 engineers across Pune and Bangalore.<br/>"
            "- Net Retention Rate (NRR): 124%. Enterprise Clients: 310 accounts.<br/>"
            "- Segment Mix: Enterprise SaaS 70%, Cloud Infrastructure 30%."
        )
        elements.append(Paragraph(pres_text, body_style))
        doc.build(elements)

    @staticmethod
    def generate_financial_model_excel(file_path: str):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "P&L"

        header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")

        headers = ["Line Item", "FY2022 (in Cr)", "FY2023 (in Cr)", "FY2024 (in Cr)", "YoY Growth %"]
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill

        rows = [
            ("Revenue", 78.00, 100.00, 125.00, 25.0),
            ("COGS", 35.00, 46.00, 56.25, 22.3),
            ("Gross Profit", 43.00, 54.00, 68.75, 27.3),
            ("Opex", 26.00, 30.00, 37.50, 25.0),
            ("EBITDA", 17.00, 24.00, 31.25, 30.2),
            ("D&A", 2.50, 3.20, 4.00, 25.0),
            ("EBIT", 14.50, 20.80, 27.25, 31.0),
            ("Interest", 2.00, 2.50, 3.25, 30.0),
            ("PBT", 12.50, 18.30, 24.00, 31.1),
            ("Tax", 3.00, 4.10, 5.25, 28.0),
            ("PAT", 9.50, 14.20, 18.75, 32.0)
        ]

        for r_idx, row in enumerate(rows, 2):
            for c_idx, val in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=val)

        ws2 = wb.create_sheet(title="Balance Sheet")
        bs_headers = ["Particulars", "FY2023", "FY2024"]
        for col_idx, h in enumerate(bs_headers, 1):
            cell = ws2.cell(row=1, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill

        bs_rows = [
            ("Total Assets", 120.00, 150.00),
            ("Total Liabilities", 52.00, 65.00),
            ("Total Debt", 30.00, 42.00),
            ("Net Worth", 68.00, 85.00),
            ("Cash & Equivalents", 14.00, 18.50)
        ]
        for r_idx, row in enumerate(bs_rows, 2):
            for c_idx, val in enumerate(row, 1):
                ws2.cell(row=r_idx, column=c_idx, value=val)

        wb.save(file_path)

    @staticmethod
    def generate_cash_flow_excel(file_path: str):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Cash Flow"

        headers = ["Cash Flow Items", "FY2023 (in Cr)", "FY2024 (in Cr)"]
        for col_idx, h in enumerate(headers, 1):
            ws.cell(row=1, column=col_idx, value=h)

        rows = [
            ("Operating Cash Flow (OCF)", 40.00, 27.20),
            ("Capital Expenditure (CapEx)", 10.00, 12.00),
            ("Free Cash Flow (FCF)", 30.00, 15.20),
            ("Financing Cash Flow", -5.00, 8.00),
            ("Net Change in Cash", 15.00, 4.50)
        ]
        for r_idx, row in enumerate(rows, 2):
            for c_idx, val in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=val)

        wb.save(file_path)
""")

# 2. Demo Seeder
write_f("demo/seeder.py", """
import os
import hashlib
from sqlalchemy.orm import Session
from app.models.models import (
    Company, Document, DocumentChunk,
    FinancialMetric, CrossDocConsistencyCheck, FinancialRisk,
    IPOReadinessScore, ReviewQueueItem, AuditLog, ProcessingStatus, MetricStatus, RiskSeverity, ReviewStatus
)
from app.demo.generate_synthetic_docs import SyntheticDocumentGenerator
from app.ingestion.pdf_parser import PDFParser
from app.ingestion.excel_parser import ExcelParser
from app.ingestion.chunking import FinancialChunker
from app.financial.metric_extractor import MetricExtractor
from app.financial.consistency_auditor import ConsistencyAuditor
from app.financial.risk_engine import RiskEngine
from app.financial.ipo_readiness_scorer import IPOReadinessScorer
from app.rag.embeddings import LocalEmbeddingEngine

class DemoSeeder:
    @classmethod
    def seed_demo_company(cls, db: Session, upload_dir: str) -> Company:
        existing = db.query(Company).filter(Company.name == "Acme Technologies Private Limited").first()
        if existing:
            return existing

        company = Company(
            name="Acme Technologies Private Limited",
            cin="U72200MH2018PTC308912",
            sector="Enterprise SaaS & Cloud Infrastructure",
            target_ipo_date="Q4 FY2025",
            description="High-growth B2B enterprise SaaS company providing cloud automation and financial document AI solutions."
        )
        db.add(company)
        db.commit()
        db.refresh(company)

        doc_paths = SyntheticDocumentGenerator.generate_all(upload_dir)
        all_metrics = []
        all_chunks = []

        for f_path in doc_paths:
            f_name = os.path.basename(f_path)
            with open(f_path, "rb") as f:
                f_hash = hashlib.sha256(f.read()).hexdigest()

            doc_type = "Annual Report" if "Annual_Report" in f_name else "Investor Presentation" if "Investor" in f_name else "Financial Model"
            f_size = os.path.getsize(f_path)

            doc = Document(
                company_id=company.id,
                filename=f_name,
                file_path=f_path,
                file_hash=f_hash,
                document_type=doc_type,
                fiscal_year="FY2024",
                page_count=2 if f_name.endswith(".pdf") else 1,
                file_size_bytes=f_size,
                processing_status=ProcessingStatus.INDEXED,
                processing_duration_ms=185
            )
            db.add(doc)
            db.commit()
            db.refresh(doc)

            if f_name.endswith(".pdf"):
                parsed = PDFParser.parse_pdf(f_path)
            else:
                parsed = ExcelParser.parse_excel(f_path)

            chunks = FinancialChunker.chunk_document(parsed, doc.id, f_name, doc_type)
            for ch in chunks:
                emb = LocalEmbeddingEngine.get_embedding(ch["chunk_text"])
                db_chunk = DocumentChunk(
                    document_id=doc.id,
                    page_number=ch.get("page_number"),
                    section_title=ch.get("section_title"),
                    chunk_index=ch.get("chunk_index"),
                    chunk_text=ch.get("chunk_text"),
                    embedding=emb,
                    token_count=ch.get("token_count"),
                    chunk_metadata=ch.get("chunk_metadata")
                )
                db.add(db_chunk)
                all_chunks.append(ch)

            extracted = MetricExtractor.extract_metrics_from_text(parsed["full_text"], f_name)
            for m in extracted:
                m_obj = FinancialMetric(
                    company_id=company.id,
                    document_id=doc.id,
                    metric_name=m["metric_name"],
                    raw_value_str=m["raw_value_str"],
                    normalized_value_inr=m["normalized_value_inr"],
                    currency=m["currency"],
                    unit=m["unit"],
                    fiscal_year=m["fiscal_year"],
                    statement_type=m["statement_type"],
                    source_document_name=m["source_document_name"],
                    source_page=m.get("source_page"),
                    confidence_score=m["confidence_score"],
                    status=MetricStatus.EXTRACTED
                )
                db.add(m_obj)
                all_metrics.append(m)

        db.commit()

        inconsistencies = ConsistencyAuditor.audit_metrics(all_metrics, company.id)
        for inc in inconsistencies:
            check_obj = CrossDocConsistencyCheck(
                company_id=company.id,
                metric_name=inc["metric_name"],
                fiscal_year=inc["fiscal_year"],
                source_a_doc_name=inc["source_a_doc_name"],
                source_a_page_or_cell=inc["source_a_page_or_cell"],
                source_a_value_raw=inc["source_a_value_raw"],
                source_a_value_normalized=inc["source_a_value_normalized"],
                source_b_doc_name=inc["source_b_doc_name"],
                source_b_page_or_cell=inc["source_b_page_or_cell"],
                source_b_value_raw=inc["source_b_value_raw"],
                source_b_value_normalized=inc["source_b_value_normalized"],
                variance_amount=inc["variance_amount"],
                variance_percentage=inc["variance_percentage"],
                severity=inc["severity"],
                status=inc["status"],
                resolution_notes=inc["resolution_notes"]
            )
            db.add(check_obj)

            queue_item = ReviewQueueItem(
                company_id=company.id,
                item_type="CONSISTENCY",
                reason=f"Variance of {inc['variance_percentage']}% in {inc['metric_name']}",
                original_payload=inc,
                review_status=ReviewStatus.PENDING
            )
            db.add(queue_item)

        risks = RiskEngine.evaluate_risks(all_metrics, inconsistencies, company.id)
        for r in risks:
            risk_obj = FinancialRisk(
                company_id=company.id,
                risk_type=r["risk_type"],
                title=r["title"],
                severity=r["severity"],
                evidence=r["evidence"],
                formula_used=r["formula_used"],
                source_citation=r["source_citation"],
                confidence_score=r["confidence_score"],
                recommended_action=r["recommended_action"]
            )
            db.add(risk_obj)

        readiness = IPOReadinessScorer.calculate_readiness(doc_paths, all_metrics, inconsistencies, risks, company.id)
        score_obj = IPOReadinessScore(
            company_id=company.id,
            overall_score=readiness["overall_score"],
            financial_completeness_score=readiness["financial_completeness_score"],
            financial_consistency_score=readiness["financial_consistency_score"],
            profitability_score=readiness["profitability_score"],
            cashflow_score=readiness["cashflow_score"],
            debt_health_score=readiness["debt_health_score"],
            growth_score=readiness["growth_score"],
            document_coverage_score=readiness["document_coverage_score"],
            breakdown_details=readiness["breakdown_details"]
        )
        db.add(score_obj)

        audit = AuditLog(
            company_id=company.id,
            action_type="WORKSPACE_INITIALIZATION",
            query_text="Automated IPO Readiness & Document Ingestion Pipeline",
            steps_executed=[
                {"step": "Document Validation & Ingestion", "status": "COMPLETED"},
                {"step": "Multi-Format Parsing (PDF & Excel)", "status": "COMPLETED"},
                {"step": "Hybrid Vector Indexing", "status": "COMPLETED"},
                {"step": "Cross-Document Consistency Audit", "status": "FLAGGED_INCONSISTENCY"},
                {"step": "10-Point Risk Engine Execution", "status": "COMPLETED"},
                {"step": "100-Point IPO Readiness Score Generation", "status": "COMPLETED"}
            ],
            latency_ms=420
        )
        db.add(audit)
        db.commit()

        print(f"Demo company seeded: {company.name}")
        return company
""")

print("Demo and synthetic generator modules created.")
